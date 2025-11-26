"""
Data export functions for Airtable records.
"""
import csv
import json
from io import StringIO, BytesIO
from typing import List, Dict, Any, Optional
from datetime import datetime


def export_to_json(
    records: List[Dict[str, Any]],
    pretty: bool = True
) -> str:
    """
    Export records to JSON string.
    
    Args:
        records: List of record dicts
        pretty: Whether to format with indentation
        
    Returns:
        JSON string
    """
    if pretty:
        return json.dumps(records, indent=2, default=str)
    return json.dumps(records, default=str)


def export_to_csv(
    records: List[Dict[str, Any]],
    include_id: bool = True
) -> str:
    """
    Export records to CSV string.
    
    Args:
        records: List of record dicts with 'id' and 'fields' keys
        include_id: Whether to include record ID column
        
    Returns:
        CSV string
    """
    if not records:
        return ""
    
    output = StringIO()
    
    # Get all unique field names from all records
    all_fields = set()
    for record in records:
        if "fields" in record:
            all_fields.update(record["fields"].keys())
    
    # Build headers
    headers = []
    if include_id:
        headers.append("record_id")
    headers.extend(sorted(all_fields))
    
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    
    # Write rows
    for record in records:
        row = {}
        if include_id:
            row["record_id"] = record.get("id", "")
        
        fields = record.get("fields", {})
        for field in all_fields:
            value = fields.get(field, "")
            # Handle lists (linked records, multi-select)
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            # Handle dicts
            elif isinstance(value, dict):
                value = json.dumps(value)
            row[field] = value
        
        writer.writerow(row)
    
    return output.getvalue()


def export_to_excel(
    records: List[Dict[str, Any]],
    include_id: bool = True
) -> bytes:
    """
    Export records to Excel file (XLSX format).
    
    Args:
        records: List of record dicts with 'id' and 'fields' keys
        include_id: Whether to include record ID column
        
    Returns:
        Excel file as bytes
        
    Requires:
        openpyxl package
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    if not records:
        # Return empty workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    # Get all unique field names
    all_fields = set()
    for record in records:
        if "fields" in record:
            all_fields.update(record["fields"].keys())
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Build headers
    headers = []
    if include_id:
        headers.append("record_id")
    headers.extend(sorted(all_fields))
    
    # Write headers with formatting
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write data rows
    for row_idx, record in enumerate(records, 2):
        col_idx = 1
        
        if include_id:
            ws.cell(row=row_idx, column=col_idx, value=record.get("id", ""))
            col_idx += 1
        
        fields = record.get("fields", {})
        for field in sorted(all_fields):
            value = fields.get(field, "")
            # Handle lists
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            # Handle dicts
            elif isinstance(value, dict):
                value = json.dumps(value)
            ws.cell(row=row_idx, column=col_idx, value=value)
            col_idx += 1
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def prepare_export_data(
    records: List[Dict[str, Any]],
    flatten_linked: bool = True,
    include_metadata: bool = False
) -> List[Dict[str, Any]]:
    """
    Prepare records for export by flattening and cleaning data.
    
    Args:
        records: Raw records from Airtable
        flatten_linked: Convert linked record IDs to readable format
        include_metadata: Include createdTime, etc.
        
    Returns:
        Cleaned list of records
    """
    cleaned = []
    
    for record in records:
        cleaned_record = {
            "id": record.get("id")
        }
        
        if include_metadata:
            cleaned_record["createdTime"] = record.get("createdTime")
        
        # Copy fields
        fields = record.get("fields", {})
        cleaned_record["fields"] = {}
        
        for key, value in fields.items():
            # Handle linked records (list of IDs)
            if flatten_linked and isinstance(value, list) and all(
                isinstance(v, str) and v.startswith("rec") for v in value
            ):
                cleaned_record["fields"][key] = f"{len(value)} linked record(s)"
            else:
                cleaned_record["fields"][key] = value
        
        cleaned.append(cleaned_record)
    
    return cleaned


class ExportFormatter:
    """Format exported data with custom options."""
    
    @staticmethod
    def format_for_email(records: List[Dict[str, Any]], max_records: int = 100) -> str:
        """Format records as plain text for email."""
        if not records:
            return "No records found."
        
        output = []
        output.append(f"Total records: {len(records)}")
        output.append("")
        
        display_records = records[:max_records]
        if len(records) > max_records:
            output.append(f"Showing first {max_records} of {len(records)} records:")
        
        for idx, record in enumerate(display_records, 1):
            output.append(f"Record {idx} (ID: {record.get('id', 'N/A')})")
            fields = record.get("fields", {})
            for key, value in fields.items():
                output.append(f"  {key}: {value}")
            output.append("")
        
        return "\n".join(output)
    
    @staticmethod
    def format_for_chat(records: List[Dict[str, Any]], max_records: int = 10) -> str:
        """Format records as markdown for Google Chat."""
        if not records:
            return "*No records found.*"
        
        output = []
        output.append(f"**Found {len(records)} record(s)**\n")
        
        display_records = records[:max_records]
        
        for idx, record in enumerate(display_records, 1):
            fields = record.get("fields", {})
            # Try to find a name field
            name = (
                fields.get("Applicant Name") or
                fields.get("Name") or
                fields.get("Contractor ID") or
                record.get("id", "")
            )
            output.append(f"{idx}. **{name}**")
            
            # Show a few key fields
            for key, value in list(fields.items())[:5]:
                if key not in ["Applicant Name", "Name", "Contractor ID"]:
                    output.append(f"   • {key}: {value}")
            output.append("")
        
        if len(records) > max_records:
            output.append(f"_...and {len(records) - max_records} more._")
        
        return "\n".join(output)

